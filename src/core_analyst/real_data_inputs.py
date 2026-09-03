from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from rasterio.warp import transform_geom
from shapely.geometry import box, mapping, shape

from core_analyst.analysts.data_zone_assessment import area_weight_polygon_values

from core_analyst.shapefile_utils import iter_shapefile_features, shapefile_metadata


DATA_ZONE_ID_PREFIX = "S010"
ELDERLY_AGE_LABELS = {"65 - 69", "70 - 74", "75 - 79", "80 - 84", "85 and over"}


@dataclass(frozen=True)
class PreparedRealInputs:
    census_attributes: str | None
    data_zone_geography: str | None
    buildings: str | None
    simd: str | None
    critical_services: str | None
    manifest: str
    unavailable: list[dict[str, str]]


def prepare_real_exposure_vulnerability_inputs(
    input_dir: str | Path = "Input",
    *,
    processed_dir: str | Path | None = None,
    clip_buildings_to_glasgow_buffer: bool = True,
) -> PreparedRealInputs:
    input_dir = Path(input_dir)
    processed_dir = Path(processed_dir) if processed_dir else input_dir / "processed"
    processed_dir.mkdir(parents=True, exist_ok=True)

    unavailable: list[dict[str, str]] = []
    data_zone_dir = processed_dir / "data_zone"
    exposure_dir = processed_dir / "exposure"
    simd_dir = processed_dir / "simd"
    data_zone_dir.mkdir(parents=True, exist_ok=True)
    exposure_dir.mkdir(parents=True, exist_ok=True)
    simd_dir.mkdir(parents=True, exist_ok=True)

    census_path: Path | None = data_zone_dir / "data_zone_attributes.csv"
    try:
        census_summary = build_census_data_zone_attributes(input_dir, census_path)
    except FileNotFoundError as exc:
        unavailable.append({"dataset": "Scotland Census 2022", "reason": str(exc)})
        census_path = None
        census_summary = {"status": "unavailable", "reason": str(exc)}

    data_zone_geography = find_data_zone_boundary(input_dir)
    data_zone_summary: dict[str, Any] = {"status": "unavailable", "feature_count": 0}
    if data_zone_geography is None:
        unavailable.append(
            {
                "dataset": "Data Zone 2022 boundaries",
                "reason": "No local Data Zone 2022 polygon boundary file was found under Input.",
            }
        )
    elif census_path and data_zone_geography.suffix.lower() in {".geojson", ".json"}:
        data_zone_geography = build_enriched_data_zone_geography(
            data_zone_geography,
            census_path,
            data_zone_dir / "glasgow_data_zones_2022_enriched.geojson",
        )
        data_zone_summary = _geojson_summary(data_zone_geography)

    simd_raw = find_simd_source(input_dir)
    simd_path = None
    simd_summary = None
    if simd_raw is None:
        unavailable.append(
            {
                "dataset": "SIMD 2020v2",
                "reason": "No local SIMD 2020v2 CSV/XLSX/GIS file was found under Input.",
            }
        )
    else:
        simd_path = simd_dir / "simd_2020v2_indicators.csv"
        simd_summary = build_simd_2020v2_indicators(simd_raw, simd_path)
        data_zone_2011 = find_data_zone_2011_boundary(input_dir)
        if data_zone_2011 and data_zone_geography:
            data_zone_geography = harmonise_simd_to_2022_data_zones(
                simd_path,
                data_zone_2011,
                data_zone_geography,
                data_zone_dir / "glasgow_data_zones_2022_enriched_simd.geojson",
            )
            data_zone_summary = _geojson_summary(data_zone_geography)
            simd_summary["target_data_zone_compatible"] = True
            simd_summary["harmonisation_method"] = "area_weighted_2011_to_2022_polygon_mean"
            simd_summary["data_zone_2011_boundary"] = str(data_zone_2011)

    critical_services = input_dir / "processed" / "facilities" / "critical_services.geojson"
    if not _nonempty_geojson(critical_services):
        critical_services = None

    buildings_path = None
    try:
        buildings_path = prepare_os_openmap_buildings(
            input_dir,
            exposure_dir / "os_openmap_buildings_glasgow_1km_buffer.geojson",
            clip_to_glasgow_buffer=clip_buildings_to_glasgow_buffer,
        )
    except FileNotFoundError as exc:
        unavailable.append({"dataset": "OS OpenMap Local buildings", "reason": str(exc)})

    manifest_path = processed_dir / "real_exposure_vulnerability_inputs_manifest.json"
    manifest = {
        "processing_note": "Raw Input files are read-only; generated adapters are written under Input/processed.",
        "target_spatial_unit": "Scotland Census 2022 Data Zone",
        "prepared": {
            "census_attributes": str(census_path) if census_path else None,
            "data_zone_geography": str(data_zone_geography) if data_zone_geography else None,
            "buildings": str(buildings_path) if buildings_path else None,
            "simd": str(simd_path) if simd_path else None,
            "simd_raw": str(simd_raw) if simd_raw else None,
            "critical_services": str(critical_services) if critical_services else None,
        },
        "pipeline_readiness": _pipeline_readiness(
            census_path=census_path,
            data_zone_geography=data_zone_geography,
            buildings_path=buildings_path,
            simd=simd_path,
            simd_compatible_with_target=bool(
                simd_summary and simd_summary.get("target_data_zone_compatible") is True
            ),
            critical_services=critical_services,
        ),
        "census_summary": census_summary,
        "data_zone_summary": data_zone_summary,
        "simd_summary": simd_summary,
        "unavailable": unavailable,
    }
    manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    return PreparedRealInputs(
        census_attributes=str(census_path) if census_path else None,
        data_zone_geography=str(data_zone_geography) if data_zone_geography else None,
        buildings=str(buildings_path) if buildings_path else None,
        simd=str(simd_path) if simd_path else None,
        critical_services=str(critical_services) if critical_services else None,
        manifest=str(manifest_path),
        unavailable=unavailable,
    )


def build_enriched_data_zone_geography(
    boundary_path: str | Path,
    attributes_path: str | Path,
    output_path: str | Path,
) -> Path:
    boundary_path = Path(boundary_path)
    output_path = Path(output_path)
    with Path(attributes_path).open(encoding="utf-8", newline="") as handle:
        attributes = {row["id"]: row for row in csv.DictReader(handle)}
    payload = json.loads(boundary_path.read_text(encoding="utf-8"))
    source_features = payload.get("features", [])
    enriched_features = []
    for feature in payload["features"]:
        properties = feature.setdefault("properties", {})
        unit_id = _case_insensitive_property(properties, "id", "dzcode", "data_zone")
        if unit_id is None:
            continue
        unit_id = str(unit_id)
        if unit_id not in attributes:
            continue
        properties.update(attributes.get(unit_id, {}))
        properties["id"] = unit_id
        properties["name"] = _case_insensitive_property(properties, "name", "dzname")
        for key in ("population", "elderly_count", "elderly_prop", "occupied_households", "no_car_households", "no_car_household_prop"):
            if properties.get(key) not in (None, ""):
                properties[key] = float(properties[key])
        enriched_features.append(feature)
    payload["features"] = enriched_features
    if source_features and not enriched_features:
        raise ValueError(
            "Data Zone boundary is non-empty but no features matched Census attributes; "
            "check the Data Zone id field and release year."
        )
    payload["metadata"] = {
        **payload.get("metadata", {}),
        "source_feature_count": len(source_features),
        "matched_feature_count": len(enriched_features),
        "unmatched_feature_count": len(source_features) - len(enriched_features),
        "target_geography": "Scotland Census 2022 Data Zone",
    }
    if not payload.get("crs"):
        payload["crs"] = {"type": "name", "properties": {"name": _infer_geojson_crs(payload)}}
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload), encoding="utf-8")
    return output_path


def _pipeline_readiness(
    *,
    census_path: Path | None,
    data_zone_geography: Path | None,
    buildings_path: Path | None,
    simd: Path | None,
    simd_compatible_with_target: bool = False,
    critical_services: Path | None = None,
) -> dict[str, Any]:
    census_ready = census_path is not None and census_path.exists() and _csv_record_count(census_path) > 0
    has_geography = data_zone_geography is not None and _nonempty_geojson(data_zone_geography)
    has_simd = simd is not None
    has_usable_simd = has_simd and has_geography and simd_compatible_with_target
    return {
        "exposure": {
            "population": {
                "status": "available" if census_ready and has_geography else "unavailable",
                "uses_real_data": census_ready and has_geography,
                "reason_if_unavailable": None if census_ready and has_geography else (
                    "Census population is available as Data Zone tabular data, but Data Zone 2022 "
                    "polygon geometry is missing, so spatial population exposure is not run."
                ),
            },
            "buildings": {
                "status": "available" if buildings_path else "unavailable",
                "uses_real_data": buildings_path is not None,
                "reason_if_unavailable": None if buildings_path else "OS OpenMap Local building footprints are missing.",
            },
            "critical_infrastructure": {
                "status": "available" if critical_services else "unavailable",
                "uses_real_data": critical_services is not None,
                "reason_if_unavailable": None if critical_services else "No verified critical infrastructure dataset is present in Input.",
            },
        },
        "vulnerability": {
            "geography": {
                "status": "available" if has_geography else "unavailable",
                "uses_real_data": has_geography,
                "reason_if_unavailable": None if has_geography else (
                    "Data Zone 2022 polygon geometry is required as the vulnerability analysis geography."
                ),
            },
            "elderly_prop": {
                "status": "available" if census_ready and has_geography else "unavailable",
                "uses_real_data": census_ready and has_geography,
                "reason_if_unavailable": None if census_ready and has_geography else (
                    "Census elderly_prop is prepared, but cannot enter spatial vulnerability without Data Zone geometry."
                ),
            },
            "no_car_household_prop": {
                "status": "available" if census_ready and has_geography else "unavailable",
                "uses_real_data": census_ready and has_geography,
                "reason_if_unavailable": None if census_ready and has_geography else (
                    "Census no_car_household_prop is prepared, but cannot enter spatial vulnerability without Data Zone geometry."
                ),
            },
            "deprivation_score": {
                "status": "available" if has_usable_simd else "unavailable",
                "uses_real_data": has_usable_simd,
                "prepared_real_data": has_simd,
                "reason_if_unavailable": None if has_usable_simd else (
                    "SIMD 2020v2 indicators are prepared, but the file is native to 2011 Data Zones. "
                    "A verified 2011-to-2022 Data Zone crosswalk and Data Zone geometry are required "
                    "before deprivation_score can enter the 2022 Data Zone vulnerability pipeline."
                ),
            },
            "accessibility": {
                "status": "available" if critical_services else "unavailable",
                "uses_real_data": critical_services is not None,
                "reason_if_unavailable": None if critical_services else "No verified critical services dataset is present in Input.",
            },
        },
        "priority": {
            "status": "available" if has_geography and has_usable_simd and critical_services else "unavailable",
            "uses_real_data": bool(has_geography and has_usable_simd and critical_services),
            "reason_if_unavailable": None if has_geography and has_usable_simd and critical_services else (
                "Priority requires unit-level hazard, exposure, and vulnerability. Current building exposure is real, "
                "but Data Zone vulnerability is unavailable because Data Zone geometry and/or a compatible SIMD "
                "Data Zone source are missing."
            ),
        },
    }


def harmonise_simd_to_2022_data_zones(
    simd_csv: str | Path,
    boundary_2011: str | Path,
    target_2022: str | Path,
    output_path: str | Path,
) -> Path:
    """Area-weight SIMD indicators from 2011 polygons onto 2022 Data Zones."""

    with Path(simd_csv).open(encoding="utf-8", newline="") as handle:
        scores = {
            row["id"]: _float_or_none(row.get("deprivation_score"))
            for row in csv.DictReader(handle)
        }
    source_payload = _vector_payload(Path(boundary_2011))
    target_payload = json.loads(Path(target_2022).read_text(encoding="utf-8"))
    source_crs = _payload_crs(source_payload)
    target_crs = _payload_crs(target_payload)
    source_features = []
    for feature in source_payload.get("features", []):
        properties = feature.setdefault("properties", {})
        unit_id = _case_insensitive_property(properties, "id", "dzcode", "data_zone")
        if unit_id is None or scores.get(str(unit_id)) is None:
            continue
        geometry = feature["geometry"]
        if source_crs != target_crs:
            geometry = transform_geom(source_crs, target_crs, geometry)
        source_features.append(
            {
                "type": "Feature",
                "geometry": geometry,
                "properties": {"deprivation_score": scores[str(unit_id)]},
            }
        )
    values = area_weight_polygon_values(source_features, target_payload["features"])
    matched = 0
    for feature in target_payload["features"]:
        unit_id = str(_case_insensitive_property(feature["properties"], "id", "dzcode"))
        feature["properties"]["deprivation_score"] = values.get(unit_id)
        if values.get(unit_id) is not None:
            matched += 1
    target_payload["metadata"] = {
        **target_payload.get("metadata", {}),
        "simd_harmonisation": "area_weighted_2011_to_2022_polygon_mean",
        "simd_source_feature_count": len(source_features),
        "simd_matched_target_count": matched,
        "simd_cross_geography_approximation": True,
    }
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(target_payload), encoding="utf-8")
    return output_path


def build_simd_2020v2_indicators(source_path: str | Path, output_path: str | Path) -> dict[str, Any]:
    source_path = Path(source_path)
    output_path = Path(output_path)
    if source_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError(f"Unsupported SIMD source format: {source_path.suffix}")
    from openpyxl import load_workbook

    workbook = load_workbook(source_path, read_only=True, data_only=True)
    if "Data" not in workbook.sheetnames:
        raise ValueError("SIMD workbook does not contain a Data sheet.")
    worksheet = workbook["Data"]
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value) if value is not None else "" for value in next(rows)]
    required = {"Data_Zone", "Intermediate_Zone", "Council_area", "Income_rate", "Employment_rate"}
    missing = sorted(required - set(headers))
    if missing:
        raise ValueError(f"SIMD Data sheet is missing required columns: {missing}")

    indexes = {name: headers.index(name) for name in headers if name}
    fieldnames = [
        "id",
        "simd_data_zone_year",
        "intermediate_zone",
        "council_area",
        "income_rate",
        "employment_rate",
        "deprivation_score",
        "deprivation_score_definition",
    ]
    output_path.parent.mkdir(parents=True, exist_ok=True)
    count = 0
    missing_score = 0
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            data_zone = row[indexes["Data_Zone"]]
            if not data_zone:
                continue
            income_rate = _float_or_none(row[indexes["Income_rate"]])
            employment_rate = _float_or_none(row[indexes["Employment_rate"]])
            score = _mean_available_values([income_rate, employment_rate])
            if score is None:
                missing_score += 1
            writer.writerow(
                {
                    "id": data_zone,
                    "simd_data_zone_year": "2011",
                    "intermediate_zone": row[indexes["Intermediate_Zone"]] or "",
                    "council_area": row[indexes["Council_area"]] or "",
                    "income_rate": "" if income_rate is None else f"{income_rate:.8f}",
                    "employment_rate": "" if employment_rate is None else f"{employment_rate:.8f}",
                    "deprivation_score": "" if score is None else f"{score:.8f}",
                    "deprivation_score_definition": (
                        "Mean of SIMD 2020v2 Income_rate and Employment_rate indicators; "
                        "higher values indicate higher income/employment deprivation."
                    ),
                }
            )
            count += 1
    return {
        "source": str(source_path),
        "output": str(output_path),
        "sheet": "Data",
        "record_count": count,
        "field_count": len(headers),
        "id_field": "Data_Zone",
        "source_data_zone_year": 2011,
        "target_data_zone_year": 2022,
        "target_data_zone_compatible": False,
        "missing_deprivation_score_units": missing_score,
        "deprivation_score_definition": (
            "Mean of SIMD 2020v2 Income_rate and Employment_rate indicators; not the overall SIMD rank."
        ),
    }


def build_census_data_zone_attributes(input_dir: Path, output_path: Path) -> dict[str, Any]:
    age_path = input_dir / "Datazone2022" / "Table UV102b - Age (20) by sex.csv"
    cars_path = input_dir / "Datazone2022" / "Table UV405 - Car or van availability.csv"
    lookup_path = input_dir / "Output Area 2022 to Data Zone 2022 and Intermediate Zone 2022" / "DZ22_Lookup.csv"
    for path in (age_path, cars_path, lookup_path):
        if not path.exists():
            raise FileNotFoundError(f"Required Census input is missing: {path}")

    names = _read_data_zone_names(lookup_path)
    population: dict[str, int] = {}
    elderly: dict[str, int] = {}
    for code, categories, value in _iter_census_rows(age_path):
        if len(categories) != 2:
            continue
        age_label, sex = categories
        if sex != "All people":
            continue
        if age_label == "All people":
            population[code] = value
        if age_label in ELDERLY_AGE_LABELS:
            elderly[code] = elderly.get(code, 0) + value

    households: dict[str, int] = {}
    no_car: dict[str, int] = {}
    for code, categories, value in _iter_census_rows(cars_path):
        if categories == ["All occupied households"]:
            households[code] = value
        if categories == ["Number of cars or vans in household: No cars or vans"]:
            no_car[code] = value

    fieldnames = [
        "id",
        "name",
        "population",
        "elderly_count",
        "elderly_prop",
        "occupied_households",
        "no_car_households",
        "no_car_household_prop",
        "deprivation_score",
        "building_count",
    ]
    output_path.parent.mkdir(parents=True, exist_ok=True)
    missing_population = 0
    missing_households = 0
    missing_no_car = 0
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for code in sorted(set(names) | set(population) | set(households)):
            pop = population.get(code)
            elderly_count = elderly.get(code)
            household_total = households.get(code)
            no_car_count = no_car.get(code)
            if pop in (None, 0):
                missing_population += 1
            if household_total in (None, 0):
                missing_households += 1
            if no_car_count is None:
                missing_no_car += 1
            writer.writerow(
                {
                    "id": code,
                    "name": names.get(code, ""),
                    "population": "" if pop is None else pop,
                    "elderly_count": "" if elderly_count is None else elderly_count,
                    "elderly_prop": _ratio(elderly_count, pop),
                    "occupied_households": "" if household_total is None else household_total,
                    "no_car_households": "" if no_car_count is None else no_car_count,
                    "no_car_household_prop": _ratio(no_car_count, household_total),
                    "deprivation_score": "",
                    "building_count": "",
                }
            )
    return {
        "data_zone_count": len(set(names) | set(population) | set(households)),
        "population_records": len(population),
        "elderly_records": len(elderly),
        "household_records": len(households),
        "no_car_records": len(no_car),
        "missing_population_units": missing_population,
        "missing_household_units": missing_households,
        "missing_no_car_units": missing_no_car,
    }


def prepare_os_openmap_buildings(
    input_dir: str | Path,
    output_path: str | Path,
    *,
    clip_to_glasgow_buffer: bool = True,
) -> Path:
    input_dir = Path(input_dir)
    output_path = Path(output_path)
    if output_path.exists():
        return output_path
    building_path = next(input_dir.rglob("NS_Building.shp"), None)
    if building_path is None:
        raise FileNotFoundError("NS_Building.shp was not found under Input.")

    clip_geometry = None
    clip_bbox = None
    if clip_to_glasgow_buffer:
        glasgow_path = input_dir / "HYDROMIND_Polygon" / "HYDROMIND_Polygon" / "Glasgow_City_1km_buffer.shp"
        if glasgow_path.exists():
            glasgow_features = list(iter_shapefile_features(glasgow_path))
            if not glasgow_features:
                raise FileNotFoundError("Glasgow_City_1km_buffer.shp contains no polygon features.")
            clip_geometry = glasgow_features[0].geometry
        else:
            raise FileNotFoundError("Glasgow_City_1km_buffer.shp was not found for OS building clipping.")
        clip_bbox = box(*clip_geometry.bounds)

    features = []
    scanned = 0
    skipped_by_bbox = 0
    for feature in iter_shapefile_features(building_path):
        scanned += 1
        geometry = feature.geometry
        if clip_geometry is not None:
            if not box(*feature.bbox).intersects(clip_bbox):
                skipped_by_bbox += 1
                continue
            if not geometry.intersects(clip_geometry):
                continue
        properties = {
            "id": feature.properties.get("ID"),
            "type": "building",
            "featcode": feature.properties.get("FEATCODE"),
        }
        features.append({"type": "Feature", "geometry": mapping(geometry), "properties": properties})

    metadata = shapefile_metadata(building_path)
    geojson = {
        "type": "FeatureCollection",
        "name": "os_openmap_buildings_glasgow_1km_buffer",
        "crs": {"type": "name", "properties": {"name": metadata.crs or "EPSG:27700"}},
        "metadata": {
            "source": str(building_path),
            "source_feature_count": metadata.record_count,
            "scanned_feature_count": scanned,
            "feature_count": len(features),
            "skipped_by_bbox": skipped_by_bbox,
            "clip_geometry": "Glasgow analysis extent" if clip_geometry is not None else None,
            "geographic_unit": "building_footprint",
        },
        "features": features,
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(geojson), encoding="utf-8")
    return output_path


def find_data_zone_boundary(input_dir: str | Path) -> Path | None:
    input_dir = Path(input_dir)
    candidates = []
    for path in input_dir.rglob("*"):
        if not path.is_file():
            continue
        suffix = path.suffix.lower()
        if suffix not in {".shp", ".geojson", ".json", ".gpkg"}:
            continue
        normalized = path.name.lower().replace("_", "").replace("-", "").replace(" ", "")
        if any(token in normalized for token in ("datazone2022", "datazone", "dz2022", "dz22")):
            full_name = "".join(character.lower() for character in str(path) if character.isalnum())
            if "lookup" not in normalized and "tile" not in normalized and "2011" not in full_name:
                candidates.append(path)
    return sorted(candidates, key=lambda item: len(str(item)))[0] if candidates else None


def find_data_zone_2011_boundary(input_dir: str | Path) -> Path | None:
    input_dir = Path(input_dir)
    candidates = []
    for path in input_dir.rglob("*"):
        if not path.is_file() or path.suffix.lower() not in {".shp", ".geojson", ".json"}:
            continue
        normalized = "".join(character.lower() for character in str(path) if character.isalnum())
        if "datazone" in normalized and "2011" in normalized:
            candidates.append(path)
    return sorted(candidates, key=lambda item: len(str(item)))[0] if candidates else None


def find_simd_source(input_dir: str | Path) -> Path | None:
    input_dir = Path(input_dir)
    for path in input_dir.rglob("*"):
        if not path.is_file():
            continue
        if path.suffix.lower() not in {".csv", ".xlsx", ".xls", ".shp", ".geojson", ".json", ".gpkg"}:
            continue
        if "simd" in path.name.lower() or "deprivation" in path.name.lower():
            return path
    return None


def _iter_census_rows(path: Path) -> Iterable[tuple[str, list[str], int]]:
    with path.open(newline="", encoding="utf-8-sig", errors="replace") as handle:
        for row in csv.reader(handle):
            if len(row) < 3 or not row[0].startswith(DATA_ZONE_ID_PREFIX):
                continue
            try:
                value = int(row[-1].replace(",", ""))
            except ValueError:
                continue
            yield row[0], row[1:-1], value


def _read_data_zone_names(path: Path) -> dict[str, str]:
    with path.open(newline="", encoding="utf-8-sig", errors="replace") as handle:
        return {row["DZ22"]: row["DZ22Name"] for row in csv.DictReader(handle) if row.get("DZ22")}


def _ratio(numerator: int | None, denominator: int | None) -> str:
    if numerator is None or denominator in (None, 0):
        return ""
    return f"{numerator / denominator:.8f}"


def _float_or_none(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _mean_available_values(values: list[float | None]) -> float | None:
    finite = [value for value in values if value is not None]
    if not finite:
        return None
    return sum(finite) / len(finite)


def _case_insensitive_property(properties: dict[str, Any], *names: str) -> Any:
    normalized = {
        "".join(character.lower() for character in str(key) if character.isalnum()): value
        for key, value in properties.items()
    }
    for name in names:
        key = "".join(character.lower() for character in name if character.isalnum())
        value = normalized.get(key)
        if value not in (None, ""):
            return value
    return None


def _csv_record_count(path: Path) -> int:
    with path.open(encoding="utf-8", newline="") as handle:
        return max(sum(1 for _ in handle) - 1, 0)


def _nonempty_geojson(path: str | Path | None) -> bool:
    if path is None:
        return False
    path = Path(path)
    if not path.is_file() or path.suffix.lower() not in {".geojson", ".json"}:
        return False
    return bool(json.loads(path.read_text(encoding="utf-8")).get("features"))


def _geojson_summary(path: Path) -> dict[str, Any]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    metadata = payload.get("metadata", {})
    return {
        "status": "available" if payload.get("features") else "unavailable",
        "feature_count": len(payload.get("features", [])),
        "source_feature_count": metadata.get("source_feature_count"),
        "matched_feature_count": metadata.get("matched_feature_count"),
        "unmatched_feature_count": metadata.get("unmatched_feature_count"),
        "crs": _payload_crs(payload),
    }


def _payload_crs(payload: dict[str, Any]) -> str:
    return str(payload.get("crs", {}).get("properties", {}).get("name") or _infer_geojson_crs(payload))


def _infer_geojson_crs(payload: dict[str, Any]) -> str:
    features = payload.get("features", [])
    if not features:
        return "EPSG:4326"
    bounds = shape(features[0]["geometry"]).bounds
    return "EPSG:27700" if max(abs(value) for value in bounds) > 180 else "EPSG:4326"


def _vector_payload(path: Path) -> dict[str, Any]:
    if path.suffix.lower() in {".geojson", ".json"}:
        return json.loads(path.read_text(encoding="utf-8"))
    metadata = shapefile_metadata(path)
    return {
        "type": "FeatureCollection",
        "crs": {"type": "name", "properties": {"name": metadata.crs or "EPSG:27700"}},
        "features": [
            {
                "type": "Feature",
                "geometry": mapping(feature.geometry),
                "properties": feature.properties,
            }
            for feature in iter_shapefile_features(path)
        ],
    }
