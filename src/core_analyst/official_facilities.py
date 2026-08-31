"""Prepare reproducible official critical-service point data."""

from __future__ import annotations

import csv
import json
import re
import zipfile
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
import shapefile
from shapely.geometry import Point, mapping, shape


FACILITY_TYPES = ("hospital", "school", "care_home", "emergency_service")
POSTCODE_FIELDS = ("postcode", "post_code", "service_postcode", "pc7", "pcd", "postalcode")
EASTING_FIELDS = ("easting", "gridreferenceeasting", "ngr_e", "xcoord", "x")
NORTHING_FIELDS = ("northing", "gridreferencenorthing", "ngr_n", "ycoord", "y")
NAME_FIELDS = (
    "name", "facility_name", "hospitalname", "schoolname", "service_name",
    "servicename", "station_name", "stationname",
)


def prepare_official_facilities(
    *,
    postcode_directory: str | Path,
    facility_sources: dict[str, str | Path],
    study_area: str | Path,
    output_path: str | Path,
    quality_report_path: str | Path | None = None,
    accessibility_buffer_m: float = 20_000,
) -> dict[str, Any]:
    """Join official facility postcodes to NRS grid references and write GeoJSON."""

    invalid_types = sorted(set(facility_sources) - set(FACILITY_TYPES))
    if invalid_types:
        raise ValueError(f"Unsupported facility types: {invalid_types}")
    postcode_points = load_postcode_grid_references(postcode_directory)
    study_payload = json.loads(Path(study_area).read_text(encoding="utf-8"))
    study_geometry = shape(study_payload["features"][0]["geometry"])
    source_crs = study_payload.get("crs", {}).get("properties", {}).get("name", "EPSG:27700")
    if "27700" not in str(source_crs):
        raise ValueError("Official facility preparation expects an EPSG:27700 study area")
    accessibility_area = study_geometry.buffer(float(accessibility_buffer_m))

    features = []
    quality: dict[str, Any] = {}
    unmatched_rows = []
    for facility_type, source in facility_sources.items():
        records = list(_table_rows(Path(source)))
        matched = 0
        located = 0
        in_buffer = 0
        missing_location = 0
        for row_number, row in enumerate(records, 2):
            if facility_type == "care_home" and not _is_care_home(row):
                continue
            postcode = normalize_postcode(_field(row, POSTCODE_FIELDS))
            easting = _number(_field(row, EASTING_FIELDS))
            northing = _number(_field(row, NORTHING_FIELDS))
            point = (easting, northing) if easting is not None and northing is not None else postcode_points.get(postcode)
            if point is None:
                missing_location += 1
                if postcode:
                    unmatched_rows.append(
                        {
                            "facility_type": facility_type,
                            "row": row_number,
                            "name": _field(row, NAME_FIELDS),
                            "postcode": postcode,
                            "reason": "postcode_not_found_in_locked_nrs_directory",
                        }
                    )
                continue
            located += 1
            matched += int(easting is None)
            geometry = Point(point)
            if not accessibility_area.intersects(geometry):
                continue
            in_buffer += 1
            properties = {
                "id": f"{facility_type}-{len(features) + 1}",
                "type": facility_type,
                "name": _field(row, NAME_FIELDS) or f"{facility_type.replace('_', ' ').title()} {row_number}",
                "postcode": postcode,
                "official_source": str(source),
                "inside_study_area": study_geometry.intersects(geometry),
            }
            features.append(
                {"type": "Feature", "geometry": mapping(geometry), "properties": properties}
            )
        quality[facility_type] = {
            "source": str(source),
            "input_records": len(records),
            "postcode_matches": matched,
            "located_records": located,
            "features_within_20km_buffer": in_buffer,
            "missing_location_records": missing_location,
            "unmatched_records": sum(
                item["facility_type"] == facility_type for item in unmatched_rows
            ),
        }

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "type": "FeatureCollection",
        "name": "official_critical_services_glasgow_20km",
        "crs": {"type": "name", "properties": {"name": "EPSG:27700"}},
        "metadata": {
            "feature_count": len(features),
            "facility_types": sorted({feature["properties"]["type"] for feature in features}),
            "postcode_directory": str(postcode_directory),
            "accessibility_buffer_m": accessibility_buffer_m,
            "geocoding_method": "exact normalized postcode join to NRS grid reference",
            "online_geocoding_fallback": False,
        },
        "features": features,
    }
    output_path.write_text(json.dumps(payload), encoding="utf-8")
    report_path = Path(quality_report_path) if quality_report_path else output_path.with_name("facility_data_quality.json")
    report = {
        "status": "available" if features else "unavailable",
        "output": str(output_path),
        "feature_count": len(features),
        "by_type": quality,
        "unmatched": unmatched_rows,
    }
    report_path.write_text(json.dumps(report, indent=2), encoding="utf-8")
    return {**report, "quality_report": str(report_path)}


def load_postcode_grid_references(path: str | Path) -> dict[str, tuple[float, float]]:
    points = {}
    path = Path(path)
    if path.suffix.lower() == ".zip":
        with zipfile.ZipFile(path) as archive:
            shapefile_member = next(
                (name for name in archive.namelist() if name.lower().endswith(".shp")),
                None,
            )
            if shapefile_member:
                root = path.parent / f".{path.stem}-extracted"
                archive.extractall(root)
                shapefile_path = root / shapefile_member
                with shapefile.Reader(str(shapefile_path)) as reader:
                    fields = [field[0] for field in reader.fields[1:]]
                    for item in reader.iterShapeRecords():
                        row = {_normalise_key(key): value for key, value in zip(fields, item.record)}
                        postcode = normalize_postcode(_field(row, POSTCODE_FIELDS))
                        if postcode and item.shape.points:
                            x, y = item.shape.points[0]
                            points[postcode] = (float(x), float(y))
                if points:
                    return points
    for row in _table_rows(Path(path)):
        postcode = normalize_postcode(_field(row, POSTCODE_FIELDS))
        easting = _number(_field(row, EASTING_FIELDS))
        northing = _number(_field(row, NORTHING_FIELDS))
        if postcode and easting is not None and northing is not None:
            points[postcode] = (easting, northing)
    if not points:
        raise ValueError("NRS postcode directory contains no usable postcode grid references")
    return points


def normalize_postcode(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "")).upper()


def _table_rows(path: Path) -> Iterable[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".zip":
        with zipfile.ZipFile(path) as archive:
            members = [name for name in archive.namelist() if name.lower().endswith((".csv", ".xlsx"))]
            if not members:
                raise ValueError(f"No CSV/XLSX table found in {path}")
            root = path.parent / f".{path.stem}-extracted"
            extracted = [Path(archive.extract(member, root)) for member in members]
        for table in extracted:
            yield from _table_rows(table)
        return
    if suffix == ".csv":
        with path.open(encoding="utf-8-sig", errors="replace", newline="") as handle:
            for row in csv.DictReader(handle):
                yield {_normalise_key(key): value for key, value in row.items() if key is not None}
        return
    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(path, read_only=True, data_only=True)
        for worksheet in workbook.worksheets:
            rows = list(worksheet.iter_rows(values_only=True))
            header_index = next(
                (
                    index
                    for index, values in enumerate(rows[:25])
                    if any(
                        _normalise_key(value) in {
                            *(_normalise_key(alias) for alias in POSTCODE_FIELDS),
                            *(_normalise_key(alias) for alias in EASTING_FIELDS),
                        }
                        for value in values
                    )
                ),
                None,
            )
            if header_index is None:
                continue
            headers = [_normalise_key(value) for value in rows[header_index]]
            for values in rows[header_index + 1 :]:
                if any(value not in (None, "") for value in values):
                    yield {headers[index]: value for index, value in enumerate(values) if index < len(headers)}
        return
    raise ValueError(f"Unsupported official facility table: {path}")


def _field(row: dict[str, Any], aliases: tuple[str, ...]) -> Any:
    for alias in aliases:
        value = row.get(_normalise_key(alias))
        if value not in (None, ""):
            return value
    return None


def _normalise_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


def _number(value: Any) -> float | None:
    try:
        return None if value in (None, "") else float(value)
    except (TypeError, ValueError):
        return None


def _is_care_home(row: dict[str, Any]) -> bool:
    service_type = str(
        _field(row, ("care_service", "service_type", "servicetype", "service_type_description", "servicecategory"))
        or ""
    ).lower()
    return "care home" in service_type
